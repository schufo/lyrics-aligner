import datetime
import os
import pickle
import sys
import time
from itertools import zip_longest
from typing import Tuple, List
import ass
import textgrid
import csv
import openpyxl

from make_word2phoneme_dict import create_word2phoneme_file


def create_tsv_file(data, file_path):
    """Creates a TSV file from a list of tuples (text, start_time, end_time)"""
    with open(file_path, 'w', newline='') as file:
        writer = csv.writer(file, delimiter='\t')

        # Write the header
        writer.writerow(['Text', 'Start Time', 'End Time'])

        # Write the data rows
        for row in data:
            writer.writerow(row)


def create_excel_file(words1, words2, file_path):
    """Creates an Excel file from two lists of words"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write the words from list 1 to the first column
    for i, word in enumerate(words1, start=1):
        sheet.cell(row=i, column=1).value = word

    # Write the words from list 2 to the second column
    for i, word in enumerate(words2, start=1):
        sheet.cell(row=i, column=2).value = word

    # Save the workbook to the specified file path
    workbook.save(file_path)


def clean_words(s: str) -> str:
    """Strips punctuations from words"""
    s = s.replace(".", "").replace("…", "").replace(";", "").replace("?", "").replace(",", "") \
        .replace("!", "").replace(":", "").strip("\"") \
        .strip("(").strip(")").replace("–", "")
    return s.lower()


def read_lyrics_file(file_path):
    """Reads lyrics from a file and converts them to a list of words"""
    words = []

    with open(file_path, 'r') as file:
        for line in file:
            # Split the line into individual words
            line_words = line.strip().split()

            # Clean each word and add it to the list
            cleaned_words = [clean_words(word) for word in line_words]
            words.extend(cleaned_words)

    return words


def sec2srttime(sec: float) -> str:
    """Converts float(sec) into a hh:mm:ss,mss format"""
    m, s = divmod(sec, 60)
    h, m = divmod(m, 60)
    h, m, s = int(h), int(m), int(s)
    ms = int((sec % max(1, int(sec))) * 1000)
    hours = str(h).zfill(2)
    min = str(m).zfill(2)
    sec = str(s).zfill(2)
    millisec = str(ms).zfill(3)
    return f"{hours}:{min}:{sec},{millisec}"


def grouper(n, iterable, fillvalue=None):
    """Iterating a python iterable in a group of 'n'. Used to add multiple words to a single srt line"""
    "grouper(3, 'ABCDEFG', 'x') --> ABC DEF Gxx"
    args = [iter(iterable)] * n
    return zip_longest(fillvalue=fillvalue, *args)


def create_srt_line(*args):
    """Combining multiple tuples(can be None) into a single line of srt string"""
    out = ""
    for arg in args:
        if arg:
            out += arg[0]
            out += " "
    return out.strip()


def print_timed_lyrics(_word_timestamps):
    time_counter = 0
    for w, start_time, end_time in _word_timestamps:
        time_to_wait = max(0, start_time - time_counter)
        time.sleep(time_to_wait)
        convert = str(datetime.timedelta(seconds=start_time))
        print(f"[{w}] -------> {convert}")
        sys.stdout.flush()
        time_counter = start_time


def print_timed_subtitles(_word_timestamps):
    time_counter = 0
    for t1, t2, t3, t4 in grouper(4, _word_timestamps):
        start_time = t1[1]
        time_to_wait = max(0, start_time - time_counter)
        time.sleep(time_to_wait)
        convert = str(datetime.timedelta(seconds=start_time))
        print(f"[{create_srt_line(t1, t2, t3, t4)}] -------> {convert}")
        sys.stdout.flush()
        time_counter = start_time


def resolve_conflicts(sorted_word_tuples: List[Tuple[str, float, float]]):
    # Unpack the first word into separate variables
    text1, start_time1, end_time1 = sorted_word_tuples[0]

    # Initialize the list with the first word
    resolved_word_tuples = [(text1, start_time1, end_time1)]

    for current_word in sorted_word_tuples[1:]:
        # Get the last word in the resolved list
        last_word_text, last_word_start, last_word_end = resolved_word_tuples[-1]

        # Unpack the current word into separate variables
        current_word_text, current_word_start, current_word_end = current_word

        # If the start time of the current word is less than or equal to
        # the end time of the last word plus 0.1
        if current_word_start <= last_word_end:
            # Adjust the start time of the current word
            new_start_time = last_word_end + 0.1
            print(f"{current_word_start, last_word_end} -> {new_start_time} | Adjusting the start_time of {current_word_text} because of conflict with {last_word_text}")

            # If new duration is less than or equal to 0.1, adjust the end time
            if current_word_end - new_start_time <= 0.1:
                print(f"{current_word_text} became invalid after adjustment. Increasing the endtime.")
                new_end_time = new_start_time + 0.1
            else:
                new_end_time = current_word_end

            # Append the word with the new start and end times
            resolved_word_tuples.append((current_word_text, new_start_time, new_end_time))
        else:
            # If there's no need for adjustment, simply append the current word
            resolved_word_tuples.append(current_word)

    return resolved_word_tuples


def is_compound_word(text: str):
    return not text.isspace() and len(text.split(" ")) > 1


def subdivide(text, start_time, end_time):
    words = text.split()
    num_words = len(words)
    total_time = end_time - start_time
    time_per_word = total_time / num_words
    subdivisions = []
    for i in range(num_words):
        word_start_time = round(start_time + (i * time_per_word), 2)
        word_end_time = round(word_start_time + time_per_word, 2)
        subdivisions.append((words[i], word_start_time, word_end_time))
        start_time += 0.01
    return subdivisions


def combine_textgrid_files(file_list):
    def check_conflict(word1, word2, conflict_threshold=0.01):
        # unpack the tuples
        text1, start_time1, end_time1 = word1
        text2, start_time2, end_time2 = word2

        # check if the texts are the same
        same_text = (text1 == text2)

        # check for overlapping intervals and calculate the overlap duration
        if (start_time1 <= end_time2) and (end_time1 >= start_time2):
            overlap_duration = min(end_time1, end_time2) - max(start_time1, start_time2)
        else:
            overlap_duration = 0

        # check if the overlap duration is greater than conflict_threshold
        significant_overlap = (overlap_duration > conflict_threshold)

        # check if both conditions are met
        if same_text and significant_overlap:
            # calculate the durations of the intervals
            duration1 = end_time1 - start_time1
            duration2 = end_time2 - start_time2

            # return the bigger interval along with True
            if duration1 > duration2:
                return True, word1
            else:
                return True, word2
        else:
            print("No Conflict found")
            sys.stdout.flush()
            # if no conflict, return False and None
            return False, None

    word_labels = []
    james_annotation_time = 179.0
    with open("/Users/pushkarjajoria/Downloads/Archive/word_timestamps.ass", encoding='utf_8_sig') as handle:
        ass_obj = ass.parse(handle)
    aegi_sub_lines = ass_obj.events._lines
    time_start, i = 0.0, 0
    while time_start < james_annotation_time:
        curr_line = aegi_sub_lines[i]
        word_labels.append((curr_line.text, curr_line.start.total_seconds(), curr_line.end.total_seconds()))
        time_start = aegi_sub_lines[i+1].start.total_seconds()
        i += 1
    for i, file_path in enumerate(file_list):
        tg = textgrid.TextGrid.fromFile(file_path)
        for tier in tg:
            for interval in tier:
                text = interval.mark
                start_time = interval.minTime
                end_time = interval.maxTime
                if is_compound_word(text):
                    # Equally divide the interval for each word in the compound word
                    subdivisions = subdivide(text, start_time, end_time)
                    for w, s, e in subdivisions:
                        word_labels.append((w, s, e))
                else:
                    text = text.replace(" ", "")
                    if not text:
                        print(f"Empty annotation.")
                        continue
                    word_labels.append((text, start_time, end_time))

    return word_labels


def find_triple_repeating_words(word_list):
    """Finds all triple repeating words in a sorted list"""
    triple_repeating_words = []
    double_repeating_words = []

    prev_word = None
    repeat_count = 1

    for word, start_time, end_time in word_list:
        if word == prev_word:
            repeat_count += 1
        else:
            repeat_count = 1

        if repeat_count == 2:
            double_repeating_words.append((word, start_time, end_time))

        if repeat_count == 3:
            triple_repeating_words.append((word, start_time, end_time))

        prev_word = word

    return triple_repeating_words, double_repeating_words


def combine_dicts(*dicts):
    combined = {}

    for d in dicts:
        for key, value in d.items():
            if key in combined:
                # Check if the values are the same
                if combined[key] != value:
                    print(f"Key {key} has conflicting values: {combined[key]} and {value}.")

                    # Ask the user for their choice
                    choice = input("Enter 1 to use the first value, or 2 to use the second value: ").strip()
                    while choice not in ['1', '2']:
                        print("Invalid choice. Please enter 1 or 2.")
                        choice = input("Enter 1 to use the first value, or 2 to use the second value: ").strip()

                    # Update the value based on the user's choice
                    if choice == '2':
                        combined[key] = value

            else:
                combined[key] = value

    return combined


def load_dict_from_file(filename):
    """Load a dictionary from a tab-separated file"""
    with open(filename) as f:
        lines = f.readlines()

    word2phonemes = {}
    for line in lines:
        line = line.replace('\n', '').split('\t')
        word = line[0].lower().replace('’', "'")
        phonemes = line[1]
        word2phonemes[word] = phonemes

    return word2phonemes


def update_global_dict(new_file):
    # Load existing global dictionary
    output_path = "../dataset/word2phonemeglobal.pickle"
    with open(output_path, "rb") as f:
        global_dict = pickle.load(f)

    # Load new dictionary from the provided file
    new_dict = load_dict_from_file(new_file)

    # Combine dictionaries using combine_dicts to handle conflicts
    combined_dict = combine_dicts(global_dict, new_dict)

    # Save the updated global dictionary back to the file
    with open(output_path, "wb") as f:
        pickle.dump(combined_dict, f)

    print("Done")


def create_lyrics_from_labels(words: List[str], file_path: str):
    """Creates a text file with five words per line from a list of words"""
    with open(file_path, 'w') as file:
        for i in range(0, len(words), 5):
            line = ' '.join(words[i:i+5])
            file.write(line + '\n')


if __name__ == "__main__":
    """
    Generate word to pickle from tsv file
    """
    tsv_files = ["/Users/pushkarjajoria/Desktop/Aria data prep/casta_diva/norma_transcription.tsv"]

    for tsv_file_path in tsv_files:
        aria_folder = os.path.dirname(tsv_file_path)
        output_filepath = aria_folder + "/word2phonemes.pickle"
        create_word2phoneme_file(tsv_file_path, output_filepath)
        print(f"Created new file at {output_filepath}")